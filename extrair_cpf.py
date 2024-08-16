import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from time import sleep

# Caminho para os arquivos Excel
caminho_dados_clientes = 'dados_clientes.xlsx'
caminho_planilha_fechamento = 'planilha_fechamento.xlsx'

# 1 - Entrar na planilha e extrair o CPF do cliente
planilha_clientes = openpyxl.load_workbook(caminho_dados_clientes)
pagina_clientes = planilha_clientes['Sheet1']

# Configuração do WebDriver
service = Service(ChromeDriverManager().install())
driver = webdriver.Chrome(service=service)
driver.get('https://consultcpf-devaprender.netlify.app/')

# 2 - Abrir a planilha de fechamento
try:
    planilha_fechamento = openpyxl.load_workbook(caminho_planilha_fechamento)
    pagina_fechamento = planilha_fechamento.active
except FileNotFoundError:
    planilha_fechamento = openpyxl.Workbook()
    pagina_fechamento = planilha_fechamento.active
    pagina_fechamento.title = 'Sheet1'
    pagina_fechamento.append(['Nome', 'Valor', 'CPF', 'Vencimento', 'Status', 'Data de Pagamento', 'Método de Pagamento'])

# 3 - Loop para processar cada cliente na planilha
for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha
    
    try:
        # 4 - Entrar no site e usar o CPF da planilha para pegar o status do pagamento daquele cliente
        campo_pesquisa = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//input[@id='cpfInput']"))
        )
        campo_pesquisa.clear()
        campo_pesquisa.send_keys(cpf)
        sleep(5)
        botao_pesquisar = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']"))
        )
        botao_pesquisar.click()
        sleep(4)
        status = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//span[@id='statusLabel']"))
        )
        sleep(2)
        if status.text == 'em dia':
            data_pagamento = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//p[@id='paymentDate']"))
            )
            metodo_pagamento = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//p[@id='paymentMethod']"))
            )
            sleep(2)
            data_pagamento_limpo = data_pagamento.text.split()[3]
            metodo_pagamento_limpo = metodo_pagamento.text.split()[3]
            
            pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia', data_pagamento_limpo, metodo_pagamento_limpo])
        else:
            pagina_fechamento.append([nome, valor, cpf, vencimento, 'Pendente'])
        sleep(2)
    except Exception as e:
        print(f"Erro ao processar o cliente {nome}: {e}")

# 6 - Salvar a planilha de fechamento
planilha_fechamento.save(caminho_planilha_fechamento)

# 7 - Fechar o navegador
driver.quit()
